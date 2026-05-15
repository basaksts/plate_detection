import subprocess
import sys
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
ARTIFACT_DIR = ROOT / "test_artifacts"
JUNIT_DIR = ARTIFACT_DIR / "junit_reports"
HISTORY_XLSX = ARTIFACT_DIR / "test_history.xlsx"
LATEST_SUMMARY = ARTIFACT_DIR / "latest_summary.txt"


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return

    wb = Workbook()

    ws_runs = wb.active
    ws_runs.title = "runs"
    ws_runs.append([
        "run_id",
        "timestamp",
        "total",
        "passed",
        "failed",
        "errors",
        "skipped",
        "duration_s",
        "pass_rate",
        "junit_file",
    ])

    ws_cases = wb.create_sheet("test_cases")
    ws_cases.append([
        "run_id",
        "timestamp",
        "classname",
        "test_name",
        "status",
        "duration_s",
        "message",
    ])

    wb.save(path)


def parse_junit_xml(xml_path: Path):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    test_cases = root.findall(".//testcase")

    total = len(test_cases)
    failed = 0
    errors = 0
    skipped = 0
    passed = 0
    duration = 0.0

    rows = []

    for case in test_cases:
        classname = case.attrib.get("classname", "")
        name = case.attrib.get("name", "")
        time_s = float(case.attrib.get("time", "0") or 0)
        duration += time_s

        failure_node = case.find("failure")
        error_node = case.find("error")
        skipped_node = case.find("skipped")

        message = ""

        if failure_node is not None:
            status = "failed"
            failed += 1
            message = failure_node.attrib.get("message", "")
        elif error_node is not None:
            status = "error"
            errors += 1
            message = error_node.attrib.get("message", "")
        elif skipped_node is not None:
            status = "skipped"
            skipped += 1
            message = skipped_node.attrib.get("message", "")
        else:
            status = "passed"
            passed += 1

        rows.append({
            "classname": classname,
            "test_name": name,
            "status": status,
            "duration_s": round(time_s, 4),
            "message": message,
        })

    pass_rate = (passed / total * 100) if total else 0.0

    summary = {
        "total": total,
        "passed": passed,
        "failed": failed,
        "errors": errors,
        "skipped": skipped,
        "duration_s": round(duration, 4),
        "pass_rate": round(pass_rate, 2),
    }

    return summary, rows


def append_to_excel(run_id: str, timestamp: str, summary: dict, rows: list, junit_file: Path):
    ensure_workbook(HISTORY_XLSX)

    wb = load_workbook(HISTORY_XLSX)

    ws_runs = wb["runs"]
    ws_runs.append([
        run_id,
        timestamp,
        summary["total"],
        summary["passed"],
        summary["failed"],
        summary["errors"],
        summary["skipped"],
        summary["duration_s"],
        summary["pass_rate"],
        str(junit_file.relative_to(ROOT)),
    ])

    ws_cases = wb["test_cases"]
    for row in rows:
        ws_cases.append([
            run_id,
            timestamp,
            row["classname"],
            row["test_name"],
            row["status"],
            row["duration_s"],
            row["message"],
        ])

    wb.save(HISTORY_XLSX)


def write_latest_summary(run_id: str, timestamp: str, summary: dict):
    text = f"""TEST RUN SUMMARY
Run ID      : {run_id}
Timestamp   : {timestamp}
Total       : {summary["total"]}
Passed      : {summary["passed"]}
Failed      : {summary["failed"]}
Errors      : {summary["errors"]}
Skipped     : {summary["skipped"]}
Duration(s) : {summary["duration_s"]}
Pass Rate   : {summary["pass_rate"]:.2f}%
"""

    LATEST_SUMMARY.write_text(text, encoding="utf-8")


def main():
    ARTIFACT_DIR.mkdir(exist_ok=True)
    JUNIT_DIR.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_id = f"RUN-{timestamp}-{uuid.uuid4().hex[:6]}"

    junit_file = JUNIT_DIR / f"junit_{timestamp}.xml"

    cmd = [
        sys.executable,
        "-m",
        "pytest",
        "tests",
        "-q",
        f"--junitxml={junit_file}",
    ]

    print("Running tests...")
    print(" ".join(cmd))

    result = subprocess.run(cmd, cwd=ROOT)

    if not junit_file.exists():
        print("Junit XML oluşturulamadı. Pytest çalışırken erken hata oluşmuş olabilir.")
        sys.exit(result.returncode)

    summary, rows = parse_junit_xml(junit_file)

    append_to_excel(run_id, timestamp, summary, rows, junit_file)
    write_latest_summary(run_id, timestamp, summary)

    print("\n--- TEST RUN SUMMARY ---")
    print(f"Run ID      : {run_id}")
    print(f"Total       : {summary['total']}")
    print(f"Passed      : {summary['passed']}")
    print(f"Failed      : {summary['failed']}")
    print(f"Errors      : {summary['errors']}")
    print(f"Skipped     : {summary['skipped']}")
    print(f"Pass Rate   : {summary['pass_rate']:.2f}%")
    print(f"Excel       : {HISTORY_XLSX}")
    print(f"Junit XML   : {junit_file}")

    sys.exit(result.returncode)


if __name__ == "__main__":
    main()